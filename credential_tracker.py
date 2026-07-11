"""
Credential & License Compliance Tracker
-----------------------------------------------------------------------------
Audits employee license/certification records for a healthcare organization
against verification-frequency and documentation standards drawn from real
regulatory and accreditation frameworks:

  - The Joint Commission, Human Resources chapter, HR.01.01.01 / HR.01.02.05:
    requires primary source verification (PSV) of licensure, certification,
    or registration, with risk-based ongoing monitoring frequency.
    Reference: https://www.jointcommission.org (Standards FAQ, HR chapter)

  - CMS Conditions of Participation, 42 CFR 482.12: requires organizations to
    verify current licenses, training, and certifications before staff
    provide patient care, with documented primary source confirmation
    retained in credentialing files.

  - New Jersey personnel record retention: NJ employers must retain wage and
    hour records for at least 6 years (N.J.S.A. 34:11-4.1 et seq.) and
    personnel files for at least 2 years after separation of employment.

  - EEOC recordkeeping: covered employers must retain employment records for
    at least 1 year from a qualifying personnel action (e.g., termination).

This tool is a portfolio / learning project. It does not replace legal
counsel or an organization's actual compliance program, and the retention
and verification thresholds below are simplified defaults, not legal advice.
Always confirm current requirements with the relevant licensing board, CMS,
The Joint Commission, and qualified counsel before using logic like this in
a real HR system.
-----------------------------------------------------------------------------
"""

from __future__ import annotations
import pandas as pd
from datetime import datetime, date
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

TODAY = date(2026, 7, 10)  # fixed "as of" date so results are reproducible

# Risk-based verification cadence, modeled on Joint Commission guidance:
# high-risk clinical credentials (RN, NP, MD/DO, DEA) -> monthly
# allied health / technical certifications -> quarterly
# administrative / low-risk credentials -> annually
VERIFICATION_INTERVAL_DAYS = {
    "High": 30,
    "Medium": 90,
    "Low": 365,
}

# How far in advance to flag an approaching expiration, scaled to risk tier
EXPIRATION_WARNING_DAYS = {
    "High": 60,
    "Medium": 45,
    "Low": 30,
}

ACCEPTABLE_VERIFICATION_METHODS = {
    "Primary Source - Board Portal",
    "Primary Source - CVO Report",
    "Primary Source - DEA Database",
    "Primary Source - AHIMA Portal",
}

NJ_PERSONNEL_FILE_MIN_RETENTION_YEARS = 2  # post-separation, NJ minimum


@dataclass
class Flag:
    employee_id: str
    employee_name: str
    credential_type: str
    issue: str
    severity: str  # "High", "Medium", "Low"
    detail: str


def _parse_date(value):
    if pd.isna(value) or value == "":
        return None
    return pd.to_datetime(value).date()


def load_records(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    for col in ["issue_date", "expiration_date", "last_verification_date", "separation_date"]:
        df[col] = df[col].apply(_parse_date)
    return df


def audit_records(df: pd.DataFrame) -> list[Flag]:
    """
    Runs every record through a set of discrepancy checks. Each check maps
    to a real verification/documentation requirement rather than an
    arbitrary rule, so the flag reasons are explainable in audit terms.
    """
    flags: list[Flag] = []

    for _, row in df.iterrows():
        emp_id = row["employee_id"]
        name = row["employee_name"]
        cred = row["credential_type"]
        tier = row["risk_tier"]

        # 1. Missing credential number -> can't be primary-source verified at all
        if pd.isna(row["credential_number"]) or str(row["credential_number"]).strip() == "":
            flags.append(Flag(
                emp_id, name, cred, "Missing credential number", "High",
                "No credential/license number on file. Primary source verification "
                "cannot be completed or documented without this identifier."
            ))

        # 2. Expired credential
        if row["expiration_date"] and row["expiration_date"] < TODAY:
            days_expired = (TODAY - row["expiration_date"]).days
            flags.append(Flag(
                emp_id, name, cred, "Credential expired", "High",
                f"Expired {days_expired} day(s) ago on {row['expiration_date']}. "
                "Per Joint Commission HR standards, staff should not provide "
                "services under a lapsed license/certification."
            ))

        # 3. Approaching expiration, within risk-tier-adjusted warning window
        elif row["expiration_date"]:
            days_to_expiry = (row["expiration_date"] - TODAY).days
            warning_window = EXPIRATION_WARNING_DAYS.get(tier, 30)
            if 0 <= days_to_expiry <= warning_window:
                flags.append(Flag(
                    emp_id, name, cred, "Expiration approaching", "Medium" if tier != "High" else "High",
                    f"Expires in {days_to_expiry} day(s) on {row['expiration_date']}. "
                    f"Renewal should be initiated now given the {warning_window}-day "
                    f"review window for {tier}-risk credentials."
                ))

        # 4. No documented verification at all
        if row["last_verification_date"] is None:
            flags.append(Flag(
                emp_id, name, cred, "No verification on file", "High",
                "No verification date recorded. Joint Commission HR.01.02.05 "
                "requires documented verification, including who verified, "
                "when, and by what method."
            ))
        else:
            # 5. Verification overdue for this credential's risk tier
            interval = VERIFICATION_INTERVAL_DAYS.get(tier, 365)
            days_since_verification = (TODAY - row["last_verification_date"]).days
            if days_since_verification > interval:
                flags.append(Flag(
                    emp_id, name, cred, "Verification overdue", "High" if tier == "High" else "Medium",
                    f"Last verified {days_since_verification} day(s) ago; "
                    f"{tier}-risk credentials require re-verification at "
                    f"least every {interval} days."
                ))

        # 6. Verification method doesn't meet primary source standard
        method = row.get("verification_method")
        if method and pd.notna(method) and method not in ACCEPTABLE_VERIFICATION_METHODS:
            flags.append(Flag(
                emp_id, name, cred, "Non-primary-source verification", "Medium",
                f"Recorded method '{method}' does not meet the Joint Commission's "
                "primary source verification (PSV) definition. Self-reported or "
                "copy-of-license review alone does not satisfy PSV requirements."
            ))

        # 7. Record retention check for separated employees (NJ minimum)
        if row["employment_status"] == "Terminated" and row["separation_date"]:
            years_since_separation = (TODAY - row["separation_date"]).days / 365.25
            if years_since_separation < NJ_PERSONNEL_FILE_MIN_RETENTION_YEARS:
                remaining = NJ_PERSONNEL_FILE_MIN_RETENTION_YEARS - years_since_separation
                flags.append(Flag(
                    emp_id, name, cred, "Retain record - within NJ retention window", "Low",
                    f"Separated {years_since_separation:.1f} year(s) ago. New Jersey "
                    f"requires personnel files to be retained at least "
                    f"{NJ_PERSONNEL_FILE_MIN_RETENTION_YEARS} years post-separation; "
                    f"do not purge for at least {remaining:.1f} more year(s)."
                ))

    return flags


def build_report(df: pd.DataFrame, flags: list[Flag], output_path: str):
    """
    Writes a two-tab Excel workbook: a Summary tab (counts by severity and
    issue type) and a Discrepancy Detail tab (every flagged record), styled
    to be legible in an audit/survey-readiness review.
    """
    wb = Workbook()

    # ---- Summary tab ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Credential & License Compliance Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"As of: {TODAY.isoformat()}"
    ws_summary["A2"].font = Font(italic=True)

    total_records = len(df)
    total_flags = len(flags)
    by_severity = pd.Series([f.severity for f in flags]).value_counts() if flags else pd.Series(dtype=int)

    ws_summary["A4"] = "Total credential records reviewed"
    ws_summary["B4"] = total_records
    ws_summary["A5"] = "Total discrepancies flagged"
    ws_summary["B5"] = total_flags
    ws_summary["A6"] = "High severity"
    ws_summary["B6"] = int(by_severity.get("High", 0))
    ws_summary["A7"] = "Medium severity"
    ws_summary["B7"] = int(by_severity.get("Medium", 0))
    ws_summary["A8"] = "Low severity"
    ws_summary["B8"] = int(by_severity.get("Low", 0))

    for r in range(4, 9):
        ws_summary[f"A{r}"].font = Font(bold=True)

    by_issue = pd.Series([f.issue for f in flags]).value_counts() if flags else pd.Series(dtype=int)
    ws_summary["A10"] = "Discrepancies by type"
    ws_summary["A10"].font = Font(bold=True, size=12)
    row_cursor = 11
    for issue, count in by_issue.items():
        ws_summary[f"A{row_cursor}"] = issue
        ws_summary[f"B{row_cursor}"] = int(count)
        row_cursor += 1

    for col, width in {"A": 42, "B": 14}.items():
        ws_summary.column_dimensions[col].width = width

    # ---- Detail tab ----
    ws_detail = wb.create_sheet("Discrepancy Detail")
    headers = ["Employee ID", "Employee Name", "Credential Type", "Issue", "Severity", "Detail"]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    severity_fill = {
        "High": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "Medium": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "Low": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
    }

    flags_sorted = sorted(flags, key=lambda f: {"High": 0, "Medium": 1, "Low": 2}[f.severity])
    for f in flags_sorted:
        ws_detail.append([f.employee_id, f.employee_name, f.credential_type, f.issue, f.severity, f.detail])
        row_idx = ws_detail.max_row
        fill = severity_fill.get(f.severity)
        if fill:
            for c in range(1, 7):
                ws_detail.cell(row=row_idx, column=c).fill = fill

    widths = {"A": 12, "B": 18, "C": 18, "D": 26, "E": 10, "F": 70}
    for col, width in widths.items():
        ws_detail.column_dimensions[col].width = width
    for row in ws_detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)


def main():
    df = load_records("data/sample_credentials.csv")
    flags = audit_records(df)
    build_report(df, flags, "reports/compliance_report.xlsx")

    print(f"Reviewed {len(df)} credential records as of {TODAY.isoformat()}")
    print(f"Flagged {len(flags)} discrepancies")
    by_severity = pd.Series([f.severity for f in flags]).value_counts()
    for sev in ["High", "Medium", "Low"]:
        print(f"  {sev}: {int(by_severity.get(sev, 0))}")
    print("Report written to reports/compliance_report.xlsx")


if __name__ == "__main__":
    main()
